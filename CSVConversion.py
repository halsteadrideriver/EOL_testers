'''import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF


def process_csv(input_file,  output_pdf):

    try:
        # Read CSV into DataFrame
        df = pd.read_csv(input_file, skip_blank_lines=True)
        df.dropna(how='all', axis=1, inplace=True)

        # Excel Generation
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ProcessedData"

        # Styles
        header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        default_alignment = Alignment(horizontal="center", vertical="center")

        # Excel Header
        worksheet.merge_cells("A1:{}1".format(get_column_letter(len(df.columns))))
        title_cell = worksheet.cell(row=1, column=1, value="River Test Report")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = default_alignment

        # Write Header
        for col_num, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num, value=column_name)
            cell.font = Font(bold=True, size=12)
            cell.fill = header_fill
            cell.alignment = default_alignment
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20

        # Write Data
        for row_idx, row in enumerate(df.values, start=3):
            for col_idx, value in enumerate(row, start=1):
                if pd.notna(value):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = default_alignment
                    if isinstance(value, str):
                        if value.lower() == "pass":
                            cell.fill = pass_fill
                        elif value.lower() == "fail":
                            cell.fill = fail_fill

        workbook.save(output_excel)

        # PDF Generation
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, "River Test Report", ln=True, align='C')

        # Write PDF Data
        for index, row in df.iterrows():
            row_text = " | ".join([str(item) for item in row if pd.notna(item)])
            pdf.multi_cell(0, 8, row_text)

        pdf.output(output_pdf)

        return "Success: Files generated."

    except ModuleNotFoundError as e:
        return f"Error: Missing module - {e}. Please ensure all dependencies are installed."
    except Exception as e:
        return f"Error: {e}"


# Example for standalone testing
if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python script.py <input_csv> <output_excel> <output_pdf>")
    else:
        status = process_csv(sys.argv[1], sys.argv[2], sys.argv[3])
        print(status)

    try:
        # Read CSV into DataFrame
        df = pd.read_csv(input_file, skip_blank_lines=True)
        df.dropna(how='all', axis=1, inplace=True)

        # PDF Generation
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Add report title
        title = "River Test Report"
        pdf.set_font("Arial", style='B', size=11)
        pdf.cell(200, 6, title, ln=True, align='C')
        pdf.ln(8)

        # User-defined column widths
        custom_pdf_widths = [39, 52, 14, 14, 14, 14, 14, 14, 14]
        total_width = sum(custom_pdf_widths[1:])

        # First box (rows 1 to 10)
        pdf.set_fill_color(255, 199, 206)
        pdf.set_font("Arial", style='B', size=7)

        first_box_header = " | ".join(str(val) for val in df.iloc[0].values if pd.notna(val))
        pdf.cell(sum(custom_pdf_widths), 8, first_box_header, 1, 0, 'C', True)
        pdf.ln()

        pdf.set_font("Arial", size=6)
        for row in df.values[1:10]:
            pdf.cell(custom_pdf_widths[0], 8, str(row[0]) if pd.notna(row[0]) else "", 1, 0, 'C')
            merged_text = " | ".join(str(val) for val in row[1:] if pd.notna(val))
            pdf.cell(total_width, 8, merged_text, 1, 0, 'L')
            pdf.ln()

        # Second box (remaining rows)
        if len(df) > 10:
            pdf.set_font("Arial", style='B', size=8)
            second_box_header = " | ".join(str(val) for val in df.iloc[10].values if pd.notna(val))
            pdf.cell(sum(custom_pdf_widths), 8, second_box_header, 1, 0, 'C', True)
            pdf.ln()

        pdf.set_font("Arial", size=5)
        for row in df.values[11:]:
            for col_idx, item in enumerate(row):
                width = custom_pdf_widths[col_idx] if col_idx < len(custom_pdf_widths) else 30
                pdf.cell(width, 8, str(item) if pd.notna(item) else "", 1, 0, 'C')
            pdf.ln()

            # Remove unwanted pages
        pdf_pages = pdf.page_no()
        if pdf_pages > 1:
            for page in range(pdf_pages, 1, -1):  # Start from the last page to avoid reindexing issues
                if pdf.get_y() == 15:  # Check if the page is empty (only margins)
                    pdf.pages.pop(page - 1)

        pdf.output(output_pdf)
        return "Success: PDF generated."

    except Exception as e:
        return f"Error: {e}"


# Example call for LabVIEW integration
# LabVIEW can call process_csv("path/to/input.csv", "output.xlsx", "output.pdf")

import pandas as pd
from fpdf import FPDF

def process_csv(input_file,output_pdf):
    try:
        # Read CSV into DataFrame
        df = pd.read_csv(input_file, skip_blank_lines=True)
        df.dropna(how='all', axis=1, inplace=True)
        df = df.dropna(how='all')  # Remove completely empty rows'''

import pandas as pd
from fpdf import FPDF

def process_csv(data_array, output_pdf):
    try:
        # Convert 2D array to DataFrame
        df = pd.DataFrame(data_array)
        df.dropna(how='all', axis=1, inplace=True)
        df = df.dropna(how='all')  # Remove completely empty rows
        # Remove empty cells
        df.replace('', pd.NA, inplace=True)
        df.dropna(axis=0, how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)

        # PDF Generation
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Add report title
        title = "River Test Report"
        pdf.set_font("Arial", style='B', size=11)
        pdf.cell(200, 6, title, ln=True, align='C')
        pdf.ln(7)

        # User-defined column widths
        custom_pdf_widths = [39, 52, 14, 14, 14, 14, 14, 14, 14]
        total_width = sum(custom_pdf_widths[1:])

        # First box (rows 1 to 7)
        pdf.set_fill_color(199, 244, 255)
        pdf.set_font("Arial", style='B', size=7)

        first_box_header = " | ".join(str(val) for val in df.iloc[0].values if pd.notna(val))
        pdf.cell(sum(custom_pdf_widths), 8, first_box_header, 1, 0, 'C', True)
        pdf.ln()

        pdf.set_font("Arial", size=6)
        for row in df.values[1:10]:
            if pd.notna(row).any():  # Skip rows with all empty values
                pdf.cell(custom_pdf_widths[0], 8, str(row[0]) if pd.notna(row[0]) else "", 1, 0, 'C')
                merged_text = " | ".join(str(val) for val in row[1:] if pd.notna(val))

                # Apply conditional coloring
                if "pass" in merged_text.lower():
                    pdf.set_fill_color(198, 239, 206)  # Green for Pass
                    fill = True
                elif "fail" in merged_text.lower():
                    pdf.set_fill_color(255, 199, 206)  # Red for Fail
                    fill = True
                else:
                    fill = False

                pdf.cell(total_width, 8, merged_text, 1, 0, 'L', fill)
                pdf.ln()

        # Second box (remaining rows)
        pdf.set_fill_color(199, 244, 255)
        if len(df) > 10:
            pdf.set_font("Arial", style='B', size=8)
            second_box_header = " | ".join(str(val) for val in df.iloc[7].values if pd.notna(val))
            pdf.cell(sum(custom_pdf_widths), 8, second_box_header, 1, 0, 'C', True)
            pdf.ln()

        pdf.set_font("Arial", size=5)
        for row in df.values[11:]:
            if pd.notna(row).any():  # Skip rows with all empty values
                for col_idx, item in enumerate(row):
                    width = custom_pdf_widths[col_idx] if col_idx < len(custom_pdf_widths) else 30

                    # Apply conditional coloring
                    if pd.notna(item) and isinstance(item, str):
                        if item.lower() == "pass":
                            pdf.set_fill_color(198, 239, 206)  # Green for Pass
                            fill = True
                        elif item.lower() == "fail":
                            pdf.set_fill_color(255, 199, 206)  # Red for Fail
                            fill = True
                        else:
                            fill = False
                    else:
                        fill = False

                    pdf.cell(width, 8, str(item) if pd.notna(item) else "", 1, 0, 'C', fill)
                pdf.ln()

        # Remove unwanted pages
        pdf_pages = pdf.page_no()
        if pdf_pages > 1:
            for page in range(pdf_pages, 1, -1):  # Start from the last page to avoid reindexing issues
                if pdf.get_y() == 15:  # Check if the page is empty (only margins)
                    pdf.pages.pop(page - 1)

        pdf.output(output_pdf)
        return "Success: PDF generated."

    except Exception as e:
        return f"Error: {e}"


# Example call for LabVIEW integration
#process_csv("C:\ Users\A. Halstesd Evander\Documents/VCU_BoxBuild/Info/DATA/Real_demo_11-02-2025_11-22_DUT2.csv", "C/Users/A. Halstesd Evander/Documents/VCU_BoxBuild/output.xlsx", "C/Users/A. Halstesd Evander/Documents/VCU_BoxBuild/output.pdf")

